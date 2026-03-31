"""
Org Chart Dashboard Server
--------------------------
Reads an Excel file, parses the employee hierarchy, and serves it as JSON
to the dashboard frontend.

Usage:
    1. Copy config.example.json to config.json and fill in your Excel path.
    2. Run: python server.py
    3. Open: http://localhost:5000
    4. Share with colleagues on the same network: http://YOUR-IP:5000
"""

import os
import json
import time
import threading
from collections import defaultdict

import openpyxl
from flask import Flask, jsonify, request, send_from_directory, Response

# ── Config ────────────────────────────────────────────────────────────────────

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')

def load_config():
    if not os.path.exists(CONFIG_FILE):
        print("\nERROR: config.json not found.")
        print("  Copy config.example.json → config.json and fill in your Excel path.\n")
        return None
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

# ── Flask app ─────────────────────────────────────────────────────────────────

app = Flask(__name__)

# ── Shared cache (updated by background watcher) ──────────────────────────────

_cache = {
    'employees': [],   # raw parsed rows
    'mtime':     0,    # last file modification time
}
_lock = threading.Lock()

# ── Excel parsing ─────────────────────────────────────────────────────────────

def get_excel_path():
    cfg = load_config()
    return cfg.get('excel_path', '') if cfg else ''

def get_sheet_name():
    cfg = load_config()
    return cfg.get('sheet_name', None) if cfg else None  # None = active sheet

def get_mtime():
    try:
        return os.path.getmtime(get_excel_path())
    except OSError:
        return 0

def parse_excel():
    path  = get_excel_path()
    sheet = get_sheet_name()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    col     = {h: i for i, h in enumerate(headers) if h}

    # Detect Mgr_Emp_Name_Lx columns dynamically, sorted by level number
    mgr_cols = sorted(
        [h for h in col if h.startswith('Mgr_Emp_Name_L')],
        key=lambda h: int(h.replace('Mgr_Emp_Name_L', ''))
    )

    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_name = row[col['Employee_Name']] if 'Employee_Name' in col else None
        if not raw_name:
            continue

        emp_name  = str(raw_name).strip()
        job_title = str(row[col['JobTitle']]).strip()      if 'JobTitle'       in col and row[col['JobTitle']]       else None
        location  = str(row[col['Location_Name']]).strip() if 'Location_Name'  in col and row[col['Location_Name']]  else None

        mgr_chain = []
        for mc in mgr_cols:
            val = row[col[mc]]
            mgr_chain.append(str(val).strip() if val else None)

        employees.append({
            'name':      emp_name,
            'job_title': job_title,
            'location':  location,
            'mgr_chain': mgr_chain,
        })

    return employees

# ── Hierarchy logic ───────────────────────────────────────────────────────────

def get_direct_manager(emp):
    """
    The direct manager is the last non-None entry in mgr_chain
    that is not the employee themselves.
    """
    last = None
    for name in emp['mgr_chain']:
        if name and name != emp['name']:
            last = name
    return last

def build_tree(employees, root_name):
    children_map = defaultdict(list)
    emp_lookup   = {}

    for emp in employees:
        emp_lookup[emp['name']] = emp
        mgr = get_direct_manager(emp)
        children_map[mgr].append(emp['name'])

    def recurse(name):
        emp = emp_lookup.get(name, {})
        return {
            'name':      name,
            'job_title': emp.get('job_title'),
            'location':  emp.get('location'),
            'children':  [recurse(c) for c in children_map.get(name, [])],
        }

    return recurse(root_name)

def get_all_roots(employees):
    """
    Returns all people who manage at least one other person in the dataset,
    sorted alphabetically.
    """
    emp_names    = {e['name'] for e in employees}
    children_map = defaultdict(list)

    for emp in employees:
        mgr = get_direct_manager(emp)
        if mgr:
            children_map[mgr].append(emp['name'])

    return sorted(name for name in children_map if name in emp_names)

# ── Cache reload ──────────────────────────────────────────────────────────────

def maybe_reload():
    mtime = get_mtime()
    with _lock:
        if mtime != _cache['mtime']:
            try:
                employees        = parse_excel()
                _cache['employees'] = employees
                _cache['mtime']     = mtime
                print(f'[reload] Excel reloaded ({len(employees)} employees)')
            except Exception as e:
                print(f'[error]  Failed to parse Excel: {e}')

def background_watcher():
    cfg      = load_config() or {}
    interval = cfg.get('poll_interval_seconds', 5)
    while True:
        maybe_reload()
        time.sleep(interval)

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    resp = send_from_directory(BASE_DIR, 'org_chart.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/api/roots')
def api_roots():
    maybe_reload()
    with _lock:
        roots = get_all_roots(_cache['employees'])
    return jsonify(roots)

@app.route('/api/hierarchy')
def api_hierarchy():
    root = request.args.get('root', '').strip()
    maybe_reload()
    with _lock:
        employees = list(_cache['employees'])
    if not root or not employees:
        return jsonify({'error': 'No root specified or no data loaded'}), 400
    return jsonify(build_tree(employees, root))

@app.route('/api/mtime')
def api_mtime():
    return jsonify({'mtime': get_mtime()})

@app.route('/api/export')
def api_export():
    """Returns a fully self-contained HTML snapshot (no server required to open)."""
    root = request.args.get('root', '').strip()
    maybe_reload()
    with _lock:
        employees = list(_cache['employees'])
    if not root or not employees:
        return jsonify({'error': 'No root specified or no data loaded'}), 400

    tree      = build_tree(employees, root)
    tree_json = json.dumps(tree, ensure_ascii=False)
    safe_name = root.replace('"', '').replace("'", '')

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Org Chart – {safe_name}</title>
<script>
  // Embedded data – no server needed
  const EXPORTED_TREE = {tree_json};
  const EXPORT_MODE   = true;
</script>
</head>
<body>
<script src="/org_chart_renderer.js"></script>
</body>
</html>"""

    # Instead of a separate JS file, embed the renderer inline by serving
    # the main HTML with the data pre-loaded.
    # We read org_chart.html and inject the data.
    html_path = os.path.join(BASE_DIR, 'org_chart.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Inject the tree data and export flag before </head>
    inject = f"""<script>
  var EXPORTED_TREE = {tree_json};
  var EXPORT_MODE   = true;
</script>
"""
    content = content.replace('</head>', inject + '</head>', 1)

    filename = f"org-chart-{root.split(',')[0].strip()}.html"
    return Response(
        content,
        mimetype='text/html',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    cfg = load_config()
    if not cfg:
        exit(1)

    maybe_reload()

    t = threading.Thread(target=background_watcher, daemon=True)
    t.start()

    port = cfg.get('port', 5000)
    print(f'\nOrg Chart Dashboard running at:')
    print(f'  Local:   http://localhost:{port}')
    print(f'  Network: http://<YOUR-IP-ADDRESS>:{port}')
    print(f'\nWatching: {get_excel_path()}')
    print('Press Ctrl+C to stop.\n')

    app.run(host='0.0.0.0', port=port, debug=False)
